from bs4 import BeautifulSoup
import requests
import random
from time import sleep
import os
import openpyxl

URL = 'https://www.avito.ru'
TIMEOUT = 10
FILEPATH = 'result.xlsx'

proxyPath = 'proxy.txt'
UAPath = 'UA.txt'



class AvitoParser(object):

    def __init__(self):
        self.URL = URL
        self.SetOptions()
        #self.SetTestingOptions()

    # Получаем настройки запроса у пользователя
    def SetOptions(self):

        # _ в конце переменной значит, что это переменная настройки

        # Далее: ввод настроек для терминала

        # Основные настройки
        print('Приветствуем! Эта программа - парсер авито! Давайте разберемся с настройками запроса: ')
        self.category_ = '/avtomobili' # Пока парсер работает только с автомобилями
        self.region_ = input('Введите регион в формате /region (Например: "/moskva"): ')
        self.q_ = input('Введите свой запрос (Например: "Лада Веста"): ')

        #Настройки для антибана
        print('Введите настройки (по очереди) проски, user-agent, паузы: 0 - выкл, 1 - вкл. '
              'Обратите внимание: для использования прокси и user-agent необходимы файлы с именами "proxy.txt" и "UA.txt" в папке с программой!')
        self.useProxy_ = input('Использовать прокси? ').strip() == '1'
        self.useUA_ = input('Использовать user-agent? ').strip() == '1'
        self.usePause_ = input('Использовать паузы? ').strip() == '1'

        if self.useProxy_:
            self.proxyList_ = self.GetListFromFile(proxyPath)
        else:
            self.proxyList_ = ['']
        if self.useUA_:
            self.UAList_ = self.GetListFromFile(UAPath)
        else:
            self.UAList_ = ['']
        
        # Цена
        print('Разберемся с фильтрами.')
        self.minCost_ = int(input('Введите минимальную цену(0, если без разницы): ').strip())
        self.maxCost_ = int(input('Введите максимальную цену(0, если без разницы): ').strip())

        # Далее часть только для машин!
        self.minRating_ = input('Какой должна быть у продавца минимальная оценка? 0 - без разницы: ').strip()
        self.minReviewsCount_ = input('Сколько минимум должно быть у продавца отзывов? 0 - без разницы: ').strip()
        self.maxOwners_ = input('Сколько максимально должно быть владельцев? 0, если без разницы: ').strip()
        self.condition_ = input('Битая или не битая? Введите "не битый", если вам нужна не битая. Введите 0, если вам без разницы: ').strip()
        self.transmission_ = input('Введите коробку передач(0 - без разницы, автомат/механика/робот/вариатор): ').strip()
        self.mileage_ = input('Какой пробег вас интересует, км? 0 - без разницы: ').strip()

        # Доп настройки
        a = input('Нужны ли вам дополнительные Фильтры? 0 - нет, 1 - да: ').strip() == '1'
        if a:
            self.power_ = input('Какая минимальная мощность вас интересует, л.с.? 0 - без разницы: ').strip()
            self.yearOfRelease_ = input('Какой год выпуска вас интересует? 0 - без разницы: ').strip()
            self.countOfDoors_ = input('Сколько дверей вас интересует? 0 - без разницы: ').strip()
            self.typeOfFuel_ = input('Какое топливо вас интересует? 0 - без разницы, бензин/дизель/газ/гибрид/электро: ').strip()
            self.volume_ = input('Какой объем двигателя вас интересует, л? 0 - без разницы: ').strip()
            self.drive_ = input('Какой привод вас интересует? 0 - без разницы, передний/задний/полный: ').strip()
        else:
            self.power_ = '0'
            self.yearOfRelease_ = '0'
            self.countOfDoors_ = '0'
            self.typeOfFuel_ = '0'
            self.volume_ = '0'
            self.drive_ = '0'

        self.header_ = {} #Чтобы добавлять что-либо в будущем
        self.params_ = {'q':self.q_, 'pmax':self.maxCost_, 'pmin':self.minCost_}

        self.names = []
        self.costs = []
        self.locations = []
        self.ratings = []
        self.reviewCounts = []
        self.owners = []
        self.conditions = []
        self.transmissions = []
        self.mileages = []
        self.links = []
        self.years = []
        self.doors = []
        self.fuels = []
        self.volumes = []
        self.drives = []
        self.powers = []

        input('✔ Все настройки установлены. Мы готовы.')


    # Возвращает Супчик, обрабатывает ошибки при подключении
    def TryToConnect(self, url, params={}, proxyList=None, UAList=None, tries=1):

        if tries > 5:
            s = BeautifulSoup()
            s.text = 'Произошла ошибка!'
            return s

        print('.')

        if self.usePause_:
            sleep(random.uniform(3,6))

        proxy = {}
        if proxyList != None:
            proxy['http'] = random.choice(proxyList)
            proxy['https'] = proxy['http']

        header = self.header_
        if UAList != None:
           header['user-agent'] = random.choice(UAList)

        try:
            r = requests.get(url, params=params, headers=header, proxies=proxy, timeout = TIMEOUT)

            if r.status_code != 200:
                print(f'❌ Ошибка {r.status_code}, пробуем еще раз...')
                return self.TryToConnect(url=url, params=params, proxyList = proxyList, UAList = UAList, tries = tries+1)
            else:
                return BeautifulSoup(r.text, 'html.parser')
        except Exception as e:
            print(f'❌ Произошла некоторая ошибка. Подключаемся еще раз...')
            print(e)
            return self.TryToConnect(url=url, params=params, proxyList = proxyList, UAList = UAList, tries = tries+1)


    # получает список из файла
    def GetListFromFile(self, path):
        if os.path.exists(path):
            with open(path) as file:
                list = file.read().split('\n')
                return list
        else:
            print('❌ Не удалось найти файл', path)
            return ['']

        # Сколько страниц?
    def GetPagesCount(self, soup):
        buttons = soup.findAll('span', class_='pagination-item-1WyVp')
        return int(buttons[-2].text)

    # Получаем список ссылок на авто на странице
    def GetItemsLinks(self, soup):
        items = soup.findAll('div', itemtype='http://schema.org/Product')
        links = []
        for i in items:
            l = URL + i.find(itemprop='url')['href']
            links.append(l)
        return links

    # создает файл и настраивает его
    def CreateXLSX(self, names, costs, locations, ratings,reviewsCounts,owners,condition, transmitions, mileages, links, powers=[],drives=[],
                   years=[],fuels=[],volumes=[],doors=[]):


        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.cell(1,1).value = 'Имя'
        for i in range(len(names)):
            sheet.cell(i+2, 1).value = names[i]

        sheet.cell(1,2).value = 'Цена, р'
        for i in range(len(costs)):
            sheet.cell(i+2,2).value = costs[i]

        sheet.cell(1,3).value = 'Место'
        for i in range(len(locations)):
            sheet.cell(i+2,3).value = locations[i]

        sheet.cell(1, 4).value = 'Рейтинг'
        for i in range(len(ratings)):
            sheet.cell(i+2,4).value = ratings[i]

        sheet.cell(1, 5).value = 'Кол-во отзывов'
        for i in range(len(reviewsCounts)):
            sheet.cell(i+2,5).value = reviewsCounts[i]

        sheet.cell(1, 6).value = 'Владельцы'
        for i in range(len(owners)):
            sheet.cell(i+2,6).value = owners[i]

        sheet.cell(1, 7).value = 'Состояние'
        for i in range(len(condition)):
            sheet.cell(i+2,7).value = condition[i]

        sheet.cell(1, 8).value = 'Пробег, км'
        for i in range(len(mileages)):
            sheet.cell(i+2,8).value = mileages[i]

        sheet.cell(1,9).value = 'Коробка передач'
        for i in range(len(transmitions)):
            sheet.cell(i+2, 9).value = transmitions[i]

        sheet.cell(1,10).value = 'Ссылка'
        for i in range(len(links)):
            sheet.cell(i+2,10).value = links[i]

        if self.power_ !='0':
            sheet.cell(1, sheet.max_column + 1).value = 'Мощность, л.с.'
            for i in range(len(powers)):
                sheet.cell(i+2, sheet.max_column + 1).value = powers[i]

        if self.drive_ != '0':
            sheet.cell(1, sheet.max_column + 1).value = 'Привод'
            for i in range(len(drives)):
                sheet.cell(i+2, sheet.max_column + 1).value = drives[i]

        if self.yearOfRelease_ !='0' :
            sheet.cell(1, sheet.max_column+1).value = 'Год выпуска'
            for i in range(len(years)):
                sheet.cell(i+2, sheet.max_column + 1).value = years[i]

        if self.typeOfFuel_ !='0':
            sheet.cell(1, sheet.max_column + 1).value = 'Тип топлива'
            for i in range(len(fuels)):
                sheet.cell(i+2, sheet.max_column + 1).value = fuels[i]

        if self.volume_ !='0':
            sheet.cell(1, sheet.max_column + 1).value = 'Объем двигателя, л'
            for i in range(len(volumes)):
                sheet.cell(i+2, sheet.max_column + 1).value = volumes[i]

        if self.countOfDoors_ !='0':
            sheet.cell(1, sheet.max_column + 1).value = 'Кол-во дверей'
            for i in range(len(doors)):
                sheet.cell(i+2, sheet.max_column + 1).value = doors[i]

        wb.save(FILEPATH)
        wb.close()

    def GetName(self, soup):
        name = soup.find('span', itemprop='name')
        if name!=None:
            return name.text
        else:
            return '-'

    def GetCost(self, soup):
        cost = soup.find('span', itemprop='price')
        if cost!=None:
            cost = cost.get('content')
        else:
            cost = '-'
        return cost

    def GetAdress(self, soup):
        adress = soup.find('span', class_='item-address__string')
        if adress !=None:
            return adress.text
        else:
            return '-'

    def GetRating(self, soup):
        rating = soup.find(class_='seller-info-rating')
        if rating != None:
            text = rating.text.split()
            rev = text[0].replace(',', '.')
            newtext = text[1].strip().split()
            count = newtext[0]
            return [rev, count]
        else:
            return ['-', '-']

    def GetParameter(self, soup, parameter):
        child = soup.find('span', text=parameter)
        if child:
            p = child.parent.text.strip().replace(parameter, '')
            p = p.strip('-').strip('+').strip('>').strip('<')
            return p
        else:
            return '-'

    #Получаем инфу из тачки, записываем ее в файл
    def ParseItem(self, link):

        soup = self.TryToConnect(link, proxyList=self.proxyList_, UAList=self.UAList_)
        name = self.GetName(soup)
        location = self.GetName(soup)

        price = self.GetCost(soup)
        # if self.minCost_ != '0':
        #     if price == 'Не удалось получить информацию о цене':
        #         return
        #     elif int(price) < int(self.minCost_):
        #         return
        # if self.minCost_ != '0':
        #     if price == 'Не удалось получить информацию о цене':
        #         return
        #     elif int(price) > int(self.maxCost_):
        #         return

        r = self.GetRating(soup)
        rating = r[0]
        reviewsCount = r[1]
        if self.minRating_ != '0':
            if rating == '-':
                return
            if float(rating) < float(self.minRating_.strip().replace(',', '.')):
                return
        if self.minReviewsCount_ !='0':
            if reviewsCount == '-':
                return
            if int(reviewsCount) < int(self.minReviewsCount_.strip()):
                return

        owners = self.GetParameter(soup, 'Владельцев по ПТС: ')
        if self.maxOwners_ != '0':
            if int(owners.strip()) > int(self.maxOwners_):
                return

        condition = self.GetParameter(soup, 'Состояние: ')
        if self.condition_ != '0':
            if condition.strip() != self.condition_:
                return

        transmition = self.GetParameter(soup, 'Коробка передач: ')
        if self.transmission_ !='0':
            if transmition != self.transmission_:
                return

        mileage = self.GetParameter(soup, 'Пробег: ')
        mileage = mileage.strip().split()[0]
        if self.mileage_ != '0':
            if int(mileage)>int(self.mileage_):
                return


        modification = self.GetParameter(soup, 'Модификация: ')
        power = modification.split()[-2].replace('(', '')
        volume = modification.split()[0]

        if self.power_ != '0':
            if int(power) < int(self.power_):
                return

        if self.volume_ != '0':
            if float(volume) != float(self.volume_):
                return

        drive = self.GetParameter(soup, 'Привод: ')
        if self.drive_ != '0':
            if drive != self.drive_:
                return

        year = self.GetParameter(soup, 'Год выпуска: ')
        if self.yearOfRelease_ != '0':
            if year != self.yearOfRelease_:
                return

        fuel = self.GetParameter(soup, 'Тип двигателя: ')
        if self.typeOfFuel_ != '0':
            if fuel != self.typeOfFuel_:
                return

        doors = self.GetParameter(soup, 'Количество дверей: ')
        if self.countOfDoors_ != '0':
            if doors != self.countOfDoors_:
                return
        print('Еще одна машина добавлена!')

        self.names.append(name)
        self.costs.append(price)
        self.locations.append(location)
        self.ratings.append(rating)
        self.reviewCounts.append(reviewsCount)
        self.owners.append(owners)
        self.conditions.append(condition)
        self.transmissions.append(transmition)
        self.mileages.append(mileage)
        self.links.append(link)
        self.powers.append(power)
        self.drives.append(drive)
        self.years.append(year)
        self.fuels.append(fuel)
        self.volumes.append(volume)
        self.doors.append(doors)








    # Основня функция
    def Parse(self):
        soup = self.TryToConnect(url=URL+self.region_+self.category_, params=self.params_, proxyList=self.proxyList_, UAList=self.UAList_)
        pagesCount = self.GetPagesCount(soup)

        #Мы подключились к первой странице, узнали, сколько всего страниц, теперь начинаем парсить каждую страницу
        pages=[]
        for i in range(pagesCount):
            print('Парсим страницу', i+1, 'из', pagesCount)
            params = self.params_
            params['p'] = i+1
            pages.append(self.TryToConnect(url=URL+self.region_+self.category_, params=params, proxyList=self.proxyList_, UAList=self.UAList_))
            linksOnThePage = self.GetItemsLinks(pages[i])

            #Парсим каждую тачку
            for l in linksOnThePage:
                self.ParseItem(l)

        self.CreateXLSX(self.names, self.costs, self.locations, self.ratings, self.reviewCounts,
                        self.owners, self.conditions, self.transmissions, self.mileages, self.links, self.powers, self.drives, self.years,
                        self.fuels, self.volumes, self.doors)
        print('✔ Парсинг успешно закончен!')




    def SetTestingOptions(self):
        self.category_ = '/avtomobili'  # Пока парсер работает только с автомобилями
        self.region_ = '/moskva'
        self.q_ = 'Lada'

        self.useProxy_ = False
        self.useUA_ = True
        self.usePause_ = True

        if self.useProxy_:
            self.proxyList_ = self.GetListFromFile(proxyPath)
        else:
            self.proxyList_ = ['']
        if self.useUA_:
            self.UAList_ = self.GetListFromFile(UAPath)
        else:
            self.UAList_ = ['']

        # Цена
        self.minCost_ = '130000'
        self.maxCost_ = '150000'

        # Далее часть только для машин!
        self.minRating_ = '4'
        self.minReviewsCount_ = '2'
        self.maxOwners_ = '2'
        self.condition_ = 'не битый'
        self.transmission_ = 'механика'
        self.mileage_ = '250000'
        self.power_ = '0'
        self.yearOfRelease_ = '0'
        self.countOfDoors_ = '0'
        self.typeOfFuel_ = '0'
        self.volume_ = '0'
        self.drive_ = '0'

        self.header_ = {}  # Чтобы добавлять что-либо в будущем
        self.params_ = {'q': self.q_, 'pmax': self.maxCost_, 'pmin': self.minCost_}

        self.names = []
        self.costs = []
        self.locations = []
        self.ratings = []
        self.reviewCounts = []
        self.owners = []
        self.conditions = []
        self.transmissions = []
        self.mileages = []
        self.links = []
        self.years = []
        self.doors = []
        self.fuels = []
        self.volumes = []
        self.drives = []
        self.powers = []

    def Test(self, soup):
        child = soup.find('span', text='Владельцев по ПТС: ')
        if child:
            print(child.parent.text.strip().replace('Владельцев по ПТС: ', ''))
        else:
            print('ytn')



















if __name__ == '__main__':
    parser = AvitoParser()
    parser.Parse()


